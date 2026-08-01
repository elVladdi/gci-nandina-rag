"""Normalize SUNAT DAM series workbooks into an auditable flat table.

The parser is designed for the SUNAT DAM layout visible in the source workbook:

- One DAM header block starts with the label ``DECLARACION :``.
- A series detail table starts with the label ``SERIE``.
- The DAM header uses repeated label/value pairs across the row.
- The detail table uses six rows of labels plus a multiline
  ``DESCRIPCION DE MERCANCIAS`` block.

The parser preserves SUNAT labels as column names. It does not translate labels
into project-specific names. It only adds one business key, ``id_unico``
(``DECLARACION`` + ``-`` + ``SERIE``), and traceability columns prefixed with
``__``.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import openpyxl

DEFAULT_INPUT = Path("data/Series - Descripciones.xlsx")
DEFAULT_OUTPUT_CSV = Path("data/interim/sunat_series_descripciones_normalized.csv")
DEFAULT_OUTPUT_XLSX = Path("data/interim/sunat_series_descripciones_normalized.xlsx")
DEFAULT_METADATA = Path("data/interim/sunat_series_descripciones_normalized_metadata.json")
DEFAULT_LABEL_AUDIT = Path("outputs/audits/sunat_series_labels_v0.1/labels.csv")
DEFAULT_DUPLICATE_AUDIT = Path("outputs/audits/sunat_series_labels_v0.1/id_unico_duplicates.csv")

DESCRIPTION_LINE_COUNT = 5
DESCRIPTION_BASE_LABEL = "DESCRIPCION DE MERCANCIAS"
DESCRIPTION_LINE_COLUMNS = [f"{DESCRIPTION_BASE_LABEL} {idx}" for idx in range(1, DESCRIPTION_LINE_COUNT + 1)]
DESCRIPTION_CONCAT_COLUMN = f"{DESCRIPTION_BASE_LABEL} CONCATENADA"
DERIVED_CODE_COLUMNS = ["NANDINA ORIGINAL", "Clase", "Partida", "Sub Partida", "NANDINA"]

BUSINESS_COLUMNS = ["id_unico"]
TECHNICAL_COLUMNS = [
    "__record_id",
    "__dam_index",
    "__series_index",
    "__source_file",
    "__sheet_name",
    "__dam_row_start",
    "__dam_row_end",
    "__detail_table_index",
    "__series_row_start",
    "__series_row_end",
    "__parse_warnings",
]

# The detail table has seven header rows. The first six rows contain labels
# whose values are stored in the six corresponding data rows. The seventh row
# names the multiline description block.
DETAIL_HEADER_ROWS = 7
DETAIL_VALUE_ROWS = 6
DESCRIPTION_HEADER_OFFSET = 6

BASE_IMPONIBLE_MARKER = "6. BASE IMPONIBLE"
LIQUIDACION_MARKER = "LIQUIDACION DEL ADEUDO"
CONCEPTO_MARKER = "CONCEPTO"
ULTIMO_DIA_PAGO_LABEL = "ULTIMO DIA DE PAGO"


@dataclass
class DamBlock:
    index: int
    sheet_name: str
    start_row: int
    end_row: int
    series_header_row: int | None
    header_fields: dict[str, str] = field(default_factory=dict)
    header_warnings: list[str] = field(default_factory=list)


@dataclass
class ParseResult:
    rows: list[dict[str, str]]
    columns: list[str]
    label_sources: dict[str, set[str]]
    label_counts: Counter[str]
    dam_count: int
    series_count: int
    skipped_dams_without_series: int
    warnings: list[str]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_label(label: str) -> str:
    label = clean_text(label)
    label = re.sub(r"\s*:\s*$", "", label)
    return label.strip()


def has_alpha(text: str) -> bool:
    return any(ch.isalpha() for ch in text)


def split_label_cell(value: Any) -> tuple[str, str] | None:
    """Return (label, inline_value) only for real SUNAT labels.

    Values such as ``00:00`` or ``003:BANCO`` are not labels because the part
    before ``:`` has no alphabetic character.
    """

    text = clean_text(value)
    if ":" not in text:
        return None
    left, right = text.split(":", 1)
    left = normalize_label(left)
    if not left or not has_alpha(left):
        return None
    return left, right.strip()


def is_declaration_row(row: list[Any]) -> bool:
    return any((split_label_cell(value) or ("", ""))[0].upper() == "DECLARACION" for value in row)


def is_series_header_row(row: list[Any]) -> bool:
    return bool(row) and clean_text(row[0]).upper() == "SERIE"


def is_series_start_row(row: list[Any]) -> bool:
    first = clean_text(row[0]) if row else ""
    return bool(re.fullmatch(r"\d+", first))


def is_blank_row(values: Iterable[Any]) -> bool:
    return all(clean_text(value) == "" for value in values)


def append_value(existing: str, new_value: str) -> str:
    if not existing:
        return new_value
    if not new_value or new_value == existing:
        return existing
    return f"{existing} || {new_value}"


def set_field(row: dict[str, str], label: str, value: str, warnings: list[str]) -> None:
    label = normalize_label(label)
    value = clean_text(value)
    if not label or not value:
        return
    previous = row.get(label, "")
    if previous and previous != value:
        warnings.append(f"valor multiple para etiqueta {label!r}")
    row[label] = append_value(previous, value)


def parse_label_value_rows(rows: list[list[Any]]) -> tuple[dict[str, str], list[str]]:
    """Parse row-wise DAM label/value pairs without inventing sublabels.

    A label cell starts a value segment. Its value is the inline text after
    ``:`` plus every following non-label cell until the next label. This keeps
    unlabeled adjacent codes inside the value of the real SUNAT label.
    """

    fields: dict[str, str] = {}
    warnings: list[str] = []
    for row in rows:
        current_label = ""
        parts: list[str] = []
        for cell in row:
            split = split_label_cell(cell)
            if split:
                if current_label:
                    set_field(fields, current_label, " ".join(parts), warnings)
                current_label, inline = split
                parts = [inline] if inline else []
            else:
                value = clean_text(cell)
                if current_label and value:
                    parts.append(value)
        if current_label:
            set_field(fields, current_label, " ".join(parts), warnings)
    return fields, warnings


def find_row(rows: list[list[Any]], start: int, end: int, marker: str) -> int | None:
    marker_upper = marker.upper()
    for row_number in range(start, end + 1):
        values = [clean_text(value).upper() for value in rows[row_number - 1]]
        if marker_upper in values:
            return row_number
    return None


def parse_base_imponible(rows: list[list[Any]], start: int, end: int) -> dict[str, str]:
    """Parse the 6. BASE IMPONIBLE matrix as row-label + column-label fields."""

    fields: dict[str, str] = {}
    base_row = find_row(rows, start, end, BASE_IMPONIBLE_MARKER)
    if base_row is None:
        return fields
    header = rows[base_row - 1]
    column_labels = {idx: clean_text(cell) for idx, cell in enumerate(header) if clean_text(cell)}
    for row_number in range(base_row + 1, end + 1):
        row = rows[row_number - 1]
        row_label = clean_text(row[0]) if row else ""
        if not row_label:
            continue
        row_label_upper = row_label.upper()
        if row_label_upper.startswith(LIQUIDACION_MARKER) or row_label_upper.startswith(CONCEPTO_MARKER):
            break
        if not row_label.startswith("6."):
            continue
        for col_idx, col_label in column_labels.items():
            if col_idx == 0 or col_idx >= len(row):
                continue
            value = clean_text(row[col_idx])
            if value:
                fields[f"{row_label} - {col_label}"] = value
    return fields


def parse_liquidacion(rows: list[list[Any]], start: int, end: int) -> dict[str, str]:
    """Parse LIQUIDACION DEL ADEUDO as concept + amount-type columns."""

    fields: dict[str, str] = {}
    concept_row = find_row(rows, start, end, CONCEPTO_MARKER)
    if concept_row is None:
        return fields
    header = rows[concept_row - 1]
    amount_columns = {idx: clean_text(cell) for idx, cell in enumerate(header) if idx > 0 and clean_text(cell)}
    for row_number in range(concept_row + 1, end + 1):
        row = rows[row_number - 1]
        concept = clean_text(row[0]) if row else ""
        if not concept:
            continue
        if concept.upper().startswith(ULTIMO_DIA_PAGO_LABEL):
            break
        for col_idx, amount_label in amount_columns.items():
            if col_idx >= len(row):
                continue
            value = clean_text(row[col_idx])
            if value:
                fields[f"{concept} - {amount_label}"] = value
    return fields


def parse_dam_header(rows: list[list[Any]], dam_start: int, dam_end: int, series_header: int | None) -> tuple[dict[str, str], list[str]]:
    header_end = (series_header - 1) if series_header else dam_end
    header_rows = rows[dam_start - 1 : header_end]
    fields, warnings = parse_label_value_rows(header_rows)
    fields.update(parse_base_imponible(rows, dam_start, header_end))
    fields.update(parse_liquidacion(rows, dam_start, header_end))
    return fields, warnings


def detail_label_maps(header_rows: list[list[Any]]) -> tuple[list[dict[int, str]], str]:
    maps: list[dict[int, str]] = []
    for offset in range(DETAIL_VALUE_ROWS):
        labels: dict[int, str] = {}
        if offset < len(header_rows):
            for col_idx, cell in enumerate(header_rows[offset]):
                label = normalize_label(clean_text(cell))
                if label:
                    labels[col_idx] = label
        maps.append(labels)

    description_label = DESCRIPTION_BASE_LABEL
    if len(header_rows) > DESCRIPTION_HEADER_OFFSET:
        for cell in header_rows[DESCRIPTION_HEADER_OFFSET]:
            label = normalize_label(clean_text(cell))
            if label:
                description_label = label
                break
    return maps, description_label


def normalize_nandina_from_label(row: dict[str, str]) -> str:
    raw = row.get("NANDINA", "")
    digits = re.sub(r"\D", "", raw)
    return digits[:8] if len(digits) >= 8 else ""


def set_derived_code_columns(row: dict[str, str], warnings: list[str]) -> None:
    """Derive tariff hierarchy columns from the visible NANDINA value.

    SUNAT often provides the code with separators, for example
    ``70.09.10.00.00``. The project needs retrieval/evaluation columns without
    dots: class/chapter (2), heading (4), subheading (6), and NANDINA (8).
    The original visible value is preserved in ``NANDINA ORIGINAL``.
    """

    raw_nandina = row.get("NANDINA", "")
    row["NANDINA ORIGINAL"] = raw_nandina
    digits = re.sub(r"\D", "", raw_nandina)
    if len(digits) < 8:
        warnings.append(f"NANDINA con menos de 8 digitos: {raw_nandina!r}")
        row["Clase"] = digits[:2]
        row["Partida"] = digits[:4]
        row["Sub Partida"] = digits[:6]
        row["NANDINA"] = digits
        return

    row["Clase"] = digits[:2]
    row["Partida"] = digits[:4]
    row["Sub Partida"] = digits[:6]
    row["NANDINA"] = digits[:8]


def parse_series_block(
    dam: DamBlock,
    all_rows: list[list[Any]],
    series_start: int,
    series_end: int,
    label_maps: list[dict[int, str]],
    description_label: str,
    source_file: Path,
    detail_table_index: int,
) -> dict[str, str]:
    warnings: list[str] = []
    series_number = clean_text(all_rows[series_start - 1][0])
    declaration = clean_text(dam.header_fields.get("DECLARACION", ""))
    id_unico = f"{declaration}-{series_number}" if declaration and series_number else ""

    record: dict[str, str] = dict(dam.header_fields)
    record.update(
        {
            "id_unico": id_unico,
            "__record_id": id_unico or f"SUNAT-DAM-{dam.index:05d}-SERIE-{series_number}",
            "__dam_index": str(dam.index),
            "__series_index": series_number,
            "__detail_table_index": str(detail_table_index),
            "__source_file": str(source_file),
            "__sheet_name": dam.sheet_name,
            "__dam_row_start": str(dam.start_row),
            "__dam_row_end": str(dam.end_row),
            "__series_row_start": str(series_start),
            "__series_row_end": str(series_end),
        }
    )

    for offset, label_by_col in enumerate(label_maps):
        row_number = series_start + offset
        if row_number > series_end or row_number > len(all_rows):
            continue
        values = all_rows[row_number - 1]
        for col_idx, label in label_by_col.items():
            if col_idx < len(values):
                set_field(record, label, clean_text(values[col_idx]), warnings)

    description_lines: list[str] = []
    for row_number in range(series_start + DETAIL_VALUE_ROWS, series_end + 1):
        if row_number > len(all_rows):
            continue
        values = all_rows[row_number - 1]
        if is_blank_row(values):
            continue
        line_parts = [clean_text(value) for value in values if clean_text(value)]
        if line_parts:
            description_lines.append(" ".join(line_parts))

    for idx, column in enumerate(DESCRIPTION_LINE_COLUMNS):
        record[column] = description_lines[idx] if idx < len(description_lines) else ""
    record[DESCRIPTION_CONCAT_COLUMN] = " ".join(description_lines)
    if len(description_lines) > DESCRIPTION_LINE_COUNT:
        warnings.append(
            f"descripcion con {len(description_lines)} lineas; se conservaron primeras {DESCRIPTION_LINE_COUNT} columnas y concatenada completa"
        )

    set_derived_code_columns(record, warnings)
    record["__parse_warnings"] = "; ".join(dam.header_warnings + warnings)
    return record


def load_sheet_rows(path: Path, sheet_name: str | None) -> tuple[str, list[list[Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return worksheet.title, rows


def find_dam_blocks(sheet_name: str, rows: list[list[Any]]) -> list[DamBlock]:
    declaration_rows = [idx for idx, row in enumerate(rows, start=1) if is_declaration_row(row)]
    blocks: list[DamBlock] = []
    for index, start in enumerate(declaration_rows, start=1):
        end = (declaration_rows[index] - 1) if index < len(declaration_rows) else len(rows)
        series_header = None
        for row_number in range(start, end + 1):
            if is_series_header_row(rows[row_number - 1]):
                series_header = row_number
                break
        header_fields, header_warnings = parse_dam_header(rows, start, end, series_header)
        blocks.append(
            DamBlock(
                index=index,
                sheet_name=sheet_name,
                start_row=start,
                end_row=end,
                series_header_row=series_header,
                header_fields=header_fields,
                header_warnings=header_warnings,
            )
        )
    return blocks


def parse_workbook(path: Path, sheet_name: str | None = None, limit_dams: int | None = None) -> ParseResult:
    sheet, rows = load_sheet_rows(path, sheet_name)
    dams = find_dam_blocks(sheet, rows)
    if limit_dams is not None:
        dams = dams[:limit_dams]

    normalized_rows: list[dict[str, str]] = []
    label_sources: dict[str, set[str]] = defaultdict(set)
    label_counts: Counter[str] = Counter()
    warnings: list[str] = []
    skipped = 0

    for dam in dams:
        for label in dam.header_fields:
            label_sources[label].add("dam_header")
        if dam.series_header_row is None:
            skipped += 1
            warnings.append(f"DAM {dam.index} sin tabla SERIE")
            continue

        table_headers = [
            row_number
            for row_number in range(dam.series_header_row, dam.end_row + 1)
            if is_series_header_row(rows[row_number - 1])
        ]
        for table_pos, table_header in enumerate(table_headers, start=1):
            table_end = (table_headers[table_pos] - 1) if table_pos < len(table_headers) else dam.end_row
            header_rows = rows[table_header - 1 : table_header - 1 + DETAIL_HEADER_ROWS]
            label_maps, description_label = detail_label_maps(header_rows)
            for label_map in label_maps:
                for label in label_map.values():
                    label_sources[label].add("series_detail")
            for column in DESCRIPTION_LINE_COLUMNS + [DESCRIPTION_CONCAT_COLUMN]:
                label_sources[column].add("series_description")
            for column in DERIVED_CODE_COLUMNS:
                label_sources[column].add("derived_from_nandina")

            search_start = table_header + DETAIL_HEADER_ROWS
            starts = [
                row_number
                for row_number in range(search_start, table_end + 1)
                if is_series_start_row(rows[row_number - 1])
            ]
            for pos, series_start in enumerate(starts):
                series_end = (starts[pos + 1] - 1) if pos + 1 < len(starts) else table_end
                parsed = parse_series_block(
                    dam,
                    rows,
                    series_start,
                    series_end,
                    label_maps,
                    description_label,
                    path,
                    table_pos,
                )
                normalized_rows.append(parsed)
                for label, value in parsed.items():
                    if label not in TECHNICAL_COLUMNS and label not in BUSINESS_COLUMNS and value:
                        label_counts[label] += 1

    sunat_columns = sorted(
        {
            label
            for row in normalized_rows
            for label in row
            if not label.startswith("__") and label not in BUSINESS_COLUMNS
        }
    )
    preferred_first = [
        "DECLARACION",
        "FECHA NUMERACION",
        "SUJETO A",
        "SERIE",
        "Clase",
        "Partida",
        "Sub Partida",
        "NANDINA",
        "DESCRIPCION DE PARTIDA ARANCELARIA",
        *DESCRIPTION_LINE_COLUMNS,
        DESCRIPTION_CONCAT_COLUMN,
    ]
    ordered_sunat = [label for label in preferred_first if label in sunat_columns]
    ordered_sunat += [label for label in sunat_columns if label not in set(ordered_sunat)]
    columns = BUSINESS_COLUMNS + ordered_sunat + TECHNICAL_COLUMNS
    return ParseResult(
        rows=normalized_rows,
        columns=columns,
        label_sources=label_sources,
        label_counts=label_counts,
        dam_count=len(dams),
        series_count=len(normalized_rows),
        skipped_dams_without_series=skipped,
        warnings=warnings,
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def write_csv(path: Path, columns: list[str], rows: list[dict[str, str]], overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"Output exists: {path}. Use --overwrite to replace it.")
    ensure_parent(path)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, columns: list[str], rows: list[dict[str, str]], overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"Output exists: {path}. Use --overwrite to replace it.")
    ensure_parent(path)
    if path.exists() and overwrite:
        try:
            with path.open("a+b"):
                pass
        except PermissionError as exc:
            raise PermissionError(f"Output appears locked by another application: {path}") from exc
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet("sunat_normalized")
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]
    workbook.save(path)


def write_label_audit(path: Path, result: ParseResult, overwrite: bool) -> None:
    rows = []
    for label in sorted(result.label_sources):
        rows.append(
            {
                "label": label,
                "sources": ",".join(sorted(result.label_sources[label])),
                "non_empty_count": str(result.label_counts.get(label, 0)),
            }
        )
    write_csv(path, ["label", "sources", "non_empty_count"], rows, overwrite)



def duplicate_payload(row: dict[str, str], columns: list[str]) -> str:
    """Stable payload used to distinguish exact duplicates from conflicts.

    Technical traceability columns are excluded because repeated source rows may
    appear in different row positions. The business key and SUNAT-derived fields
    remain in the payload.
    """

    payload = {column: row.get(column, "") for column in columns if not column.startswith("__")}
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def duplicate_audit_rows(result: ParseResult) -> list[dict[str, str]]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in result.rows:
        groups[row.get("id_unico", "")].append(row)

    audit_rows: list[dict[str, str]] = []
    for id_unico, members in sorted(groups.items()):
        if not id_unico or len(members) <= 1:
            continue
        payloads = {duplicate_payload(member, result.columns) for member in members}
        status = "duplicado_exacto" if len(payloads) == 1 else "conflicto"
        audit_rows.append(
            {
                "id_unico": id_unico,
                "total_rows": str(len(members)),
                "unique_payloads": str(len(payloads)),
                "status": status,
                "dam_indexes": ";".join(sorted({member.get("__dam_index", "") for member in members})),
                "detail_table_indexes": ";".join(sorted({member.get("__detail_table_index", "") for member in members})),
                "series_row_starts": ";".join(member.get("__series_row_start", "") for member in members),
                "nandinas": ";".join(sorted({member.get("NANDINA", "") for member in members if member.get("NANDINA", "")})),
                "descripcion_concat_muestra": members[0].get(DESCRIPTION_CONCAT_COLUMN, "")[:300],
            }
        )
    return audit_rows


def duplicate_quality_counts(result: ParseResult) -> dict[str, int]:
    rows = duplicate_audit_rows(result)
    return {
        "duplicate_id_unico_groups": len(rows),
        "duplicate_id_unico_rows_excess": sum(int(row["total_rows"]) - 1 for row in rows),
        "duplicate_id_unico_exact_groups": sum(1 for row in rows if row["status"] == "duplicado_exacto"),
        "duplicate_id_unico_conflict_groups": sum(1 for row in rows if row["status"] == "conflicto"),
    }


def write_duplicate_audit(path: Path, result: ParseResult, overwrite: bool) -> None:
    rows = duplicate_audit_rows(result)
    columns = [
        "id_unico",
        "total_rows",
        "unique_payloads",
        "status",
        "dam_indexes",
        "detail_table_indexes",
        "series_row_starts",
        "nandinas",
        "descripcion_concat_muestra",
    ]
    write_csv(path, columns, rows, overwrite)


def write_metadata(
    path: Path,
    input_path: Path,
    result: ParseResult,
    args: argparse.Namespace,
    overwrite: bool,
    xlsx_write_error: str | None = None,
) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"Metadata exists: {path}. Use --overwrite to replace it.")
    ensure_parent(path)
    nandina_valid = sum(1 for row in result.rows if re.fullmatch(r"\d{8}", normalize_nandina_from_label(row)))
    duplicate_counts = duplicate_quality_counts(result)
    metadata = {
        "input_path": str(input_path),
        "input_sha256": sha256_file(input_path),
        "sheet_name": args.sheet,
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "parser": "src.ingestion.sunat_series_parser",
        "parser_mode": "normalize_sunat_dam_series_positional_v0.2",
        "dam_blocks_detected": result.dam_count,
        "series_rows_output": result.series_count,
        "skipped_dams_without_series": result.skipped_dams_without_series,
        "business_columns": BUSINESS_COLUMNS,
        "technical_columns": TECHNICAL_COLUMNS,
        "sunat_label_columns": [column for column in result.columns if column not in BUSINESS_COLUMNS and not column.startswith("__")],
        "label_audit_path": str(args.label_audit),
        "duplicate_audit_path": str(args.duplicate_audit),
        "output_csv": str(args.output_csv),
        "output_xlsx": str(args.output_xlsx) if args.output_xlsx else None,
        "output_xlsx_error": xlsx_write_error,
        "quality_counts": {
            "rows_with_nandina_8_digits_best_effort": nandina_valid,
            "rows_with_id_unico": sum(1 for row in result.rows if row.get("id_unico")),
            "unique_id_unico": len({row.get("id_unico", "") for row in result.rows if row.get("id_unico")}),
            "rows_with_parse_warnings": sum(1 for row in result.rows if row.get("__parse_warnings")),
            "rows_with_descripcion_concat": sum(1 for row in result.rows if row.get(DESCRIPTION_CONCAT_COLUMN)),
            **duplicate_counts,
        },
        "warnings": result.warnings[:500],
        "notes": [
            "SUNAT labels are preserved as columns after removing only the trailing colon.",
            "The id_unico column is DECLARACION + '-' + SERIE.",
            "Clase, Partida, Sub Partida and NANDINA are derived from the visible SUNAT NANDINA code and stored without dots.",
            "DESCRIPCION DE MERCANCIAS is preserved as five source lines plus one concatenated search field.",
            "Repeated id_unico values are preserved and classified in the duplicate audit.",
            "Columns starting with __ are technical traceability columns added by the parser.",
            "Base imponible and liquidacion columns combine visible row labels and visible table headers.",
        ],
    }
    with path.open("w", encoding="utf-8") as handle:
        json.dump(metadata, handle, ensure_ascii=False, indent=2)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize a SUNAT DAM series workbook into one row per series, preserving real SUNAT labels as columns."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input SUNAT .xlsx workbook.")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to the first worksheet.")
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV, help="Normalized CSV output path.")
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX, help="Optional normalized XLSX output path. Use '' to skip.")
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA, help="Metadata JSON output path.")
    parser.add_argument("--label-audit", type=Path, default=DEFAULT_LABEL_AUDIT, help="Detected label audit CSV output path.")
    parser.add_argument("--duplicate-audit", type=Path, default=DEFAULT_DUPLICATE_AUDIT, help="id_unico duplicate audit CSV output path.")
    parser.add_argument("--limit-dams", type=int, default=None, help="Optional development limit for the first N DAM blocks.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing outputs.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print summary without writing outputs.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input
    if not input_path.exists():
        print(f"ERROR: input not found: {input_path}", file=sys.stderr)
        return 2

    output_xlsx = args.output_xlsx
    if output_xlsx is not None and str(output_xlsx).strip() == "":
        output_xlsx = None
        args.output_xlsx = None

    result = parse_workbook(input_path, sheet_name=args.sheet, limit_dams=args.limit_dams)
    sunat_columns_count = len([c for c in result.columns if c not in BUSINESS_COLUMNS and not c.startswith("__")])
    print(f"DAM blocks detected: {result.dam_count}")
    print(f"Series rows output: {result.series_count}")
    print(f"SUNAT labels detected: {sunat_columns_count}")
    print(f"Rows with id_unico: {sum(1 for row in result.rows if row.get('id_unico'))}")
    print(f"Rows with parse warnings: {sum(1 for row in result.rows if row.get('__parse_warnings'))}")

    if args.dry_run:
        print("Dry run: no files written.")
        return 0

    write_label_audit(args.label_audit, result, overwrite=args.overwrite)
    write_duplicate_audit(args.duplicate_audit, result, overwrite=args.overwrite)
    write_csv(args.output_csv, result.columns, result.rows, overwrite=args.overwrite)

    xlsx_write_error = None
    if output_xlsx is not None:
        try:
            write_xlsx(output_xlsx, result.columns, result.rows, overwrite=args.overwrite)
        except PermissionError as exc:
            xlsx_write_error = f"{type(exc).__name__}: {exc}"
            print(f"WARNING: XLSX output not written: {xlsx_write_error}", file=sys.stderr)
    write_metadata(args.metadata, input_path, result, args, overwrite=args.overwrite, xlsx_write_error=xlsx_write_error)

    print(f"CSV: {args.output_csv}")
    if output_xlsx is not None:
        print(f"XLSX: {output_xlsx}")
    print(f"Metadata: {args.metadata}")
    print(f"Label audit: {args.label_audit}")
    print(f"Duplicate audit: {args.duplicate_audit}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
